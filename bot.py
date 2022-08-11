#!/usr/bin/env python
# coding: utf-8

# In[ ]:


def auto_program(ticker):
    file_name = '8.5'
    file_path = '/Users/saranshpuri/Downloads/'
    your_name = 'BOT'
    import openpyxl
    wb = load_workbook(file_path + file_name + '.xlsx')
    if 'Annual' in wb.sheetnames:
        wb.remove(wb['Annual'])
        wb.save(file_path + file_name + '.xlsx')
    if 'Header' in wb.sheetnames:
        wb.remove(wb['Header'])
        wb.save(file_path + file_name + '.xlsx')
    if 'Quarterly' in wb.sheetnames:
        wb.remove(wb['Quarterly'])
        wb.save(file_path + file_name + '.xlsx')
    if 'Annual Compareable1' in wb.sheetnames:
        wb.remove(wb['Annual Compareable1'])
        wb.save(file_path + file_name + '.xlsx')
    if 'Annual Compareable2' in wb.sheetnames:
        wb.remove(wb['Annual Compareable2'])
        wb.save(file_path + file_name + '.xlsx')
    if 'Annual Compareable3' in wb.sheetnames:
        wb.remove(wb['Annual Compareable3'])
        wb.save(file_path + file_name + '.xlsx')
    if 'Annual Compareable4' in wb.sheetnames:
        wb.remove(wb['Annual Compareable4'])
        wb.save(file_path + file_name + '.xlsx')
    if 'Annual Compareable5' in wb.sheetnames:
        wb.remove(wb['Annual Compareable5'])
        wb.save(file_path + file_name + '.xlsx')
       
    
    def add_compare(ticker, file_name):
        ticker_up = ticker.upper()
        filter_set = quandl.get_table('SHARADAR/TICKERS', table='SF1',paginate = True)
        to_compare = quandl.get_table('SHARADAR/TICKERS', ticker = ticker, table='SF1').replace(['1 - Nano', '2 - Micro', '3 - Small', '4 - Mid','5 - Large','6 - Mega'], [1, 2, 3, 4, 5, 6])
        type_filter = filter_set.loc[(filter_set['sector'] == to_compare['sector'][0]) & (filter_set['industry'] == to_compare['industry'][0])]
        data_set_final = type_filter.replace(['1 - Nano', '2 - Micro', '3 - Small', '4 - Mid','5 - Large','6 - Mega'], [1, 2, 3, 4, 5, 6]).set_index('ticker')
        if to_compare['scalemarketcap'][0] == 1:
            filter1 = data_set_final.loc[(data_set_final['scalemarketcap']<=3)]
        elif to_compare['scalemarketcap'][0] == 2:
            filter1 = data_set_final.loc[(data_set_final['scalemarketcap']<=3 )]
        elif to_compare['scalemarketcap'][0] == 3:
            filter1 = data_set_final.loc[(data_set_final['scalemarketcap'] >= 2) & (data_set_final['scalemarketcap'] <= 4)]
        elif to_compare['scalemarketcap'][0] == 4:
            filter1 = data_set_final.loc[(data_set_final['scalemarketcap'] >= 3) & (data_set_final['scalemarketcap'] <= 5)]
        elif to_compare['scalemarketcap'][0] == 5:
            filter1 = data_set_final.loc[(data_set_final['scalemarketcap'] >= 4) & (data_set_final['scalemarketcap'] <= 6)]
        elif to_compare['scalemarketcap'][0] == 6:
            filter1 = data_set_final.loc[(data_set_final['scalemarketcap']>=4)]
        filter_out = filter1.loc[filter1['lastpricedate'] == to_compare['lastpricedate'][0]]
        list_companies = filter_out.index.tolist()
        if ticker_up in list_companies:
            list_companies.remove(ticker_up)
        final_list1 = list_companies[0:5]
        return final_list1
    final_list2 =  add_compare(ticker, file_name)

    
    ticker_df = ticker.upper()
    company_name = quandl.get_table('SHARADAR/TICKERS', ticker=ticker, table='SF1')['name'][0]
    today = datetime.today().strftime('%Y-%m-%d')
    header = {'label': ['Company Name', 'Ticker Symbol', 'Created By', 'Last Retrieved'],
        'value': [company_name, ticker_df, your_name, today]}
    header_df = pd.DataFrame(header)
    initial_quarterly = quandl.get_table('SHARADAR/SF1',dimension='ARQ', ticker=ticker)
    initial_quarterly['calendardate'] = pd.to_datetime(initial_quarterly['calendardate'])
    ordered_quarterly = initial_quarterly.sort_values(by='calendardate', ascending = True)
    initial_annual = quandl.get_table('SHARADAR/SF1',dimension='MRY', ticker=ticker)
    initial_annual['calendardate'] = pd.to_datetime(initial_annual['calendardate'])
    ordered_annual = initial_annual.sort_values(by='calendardate', ascending = True)
    
    book = load_workbook(file_path + file_name + '.xlsx')
    writer = pd.ExcelWriter(ticker + ' REPORT' + '.xlsx', engine='openpyxl')
    writer.book = book
    
    for i in range(len(final_list2)):
        if 'Annual Compareable' + str(i) in wb.sheetnames:
            wb.remove(wb['Annual Compareable' + str(i)])
        wb.save(file_path + file_name + '.xlsx')
        initial_annual_c = quandl.get_table('SHARADAR/SF1',dimension='MRY', ticker=final_list2[i])
        initial_annual_c['calendardate'] = pd.to_datetime(initial_annual_c['calendardate'])
        ordered_annual_c = initial_annual_c.sort_values(by=['calendardate'], ascending = True)\
        [['calendardate','ticker','revenue','ebitda', 'ebit', 'debt', 'roa', 'roe', 'price', 'marketcap', 'ev', 'equity']]
        fixed_annual_c = ordered_annual_c.swapaxes('index', 'columns', copy=True)
        idx_c = [2] + [i for i in range(len(fixed_annual_c)) if i != 2]
        annual_pd_c = fixed_annual_c.iloc[idx_c]
        fixed_annual_c.to_excel(writer,sheet_name='Annual Compareable' + str(i+1), index = True,header= False)
    fixed_quarterly = ordered_quarterly.drop(['assetsavg'], axis=1)\
        .swapaxes('index', 'columns', copy=True)
    idx = [2] + [i for i in range(len(fixed_quarterly)) if i != 2]
    fixed_annual = ordered_annual.drop(['assetsavg'], axis=1)\
        .swapaxes('index', 'columns', copy=True)
    idx = [2] + [i for i in range(len(fixed_annual)) if i != 2]
    
    
    annual_pd = fixed_annual.iloc[idx]
    quarterly_pd = fixed_quarterly.iloc[idx]
    # '/Users/saranshpuri/Downloads/' is a specific local path to file, will need to change for others
    header_df.to_excel(writer,sheet_name='Header', index = False,header= False)
    annual_pd.to_excel(writer,sheet_name='Annual', index = True,header= False)
    quarterly_pd.to_excel(writer,sheet_name='Quarterly', index = True,header= False)
    writer.save()
    file_share = open('/Users/saranshpuri/'+ ticker + ' REPORT' + '.xlsx', 'rb')
    return file_share


# In[ ]:


import logging
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters
import os
PORT = int(os.environ.get('PORT', 5000))

# Enable logging
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    level=logging.INFO)

logger = logging.getLogger(__name__)
TOKEN = '5569279566:AAF_lAT6bNBp5DuuGPu7SkpGx3cRLmw4-8U'

# Define a few command handlers. These usually take the two arguments update and
# context. Error handlers also receive the raised TelegramError object in error.
def start(update, context):
    """Send a message when the command /start is issued."""
    update.message.reply_text('Hi!')

def help(update, context):
    """Send a message when the command /help is issued."""
    update.message.reply_text('Help!')

def echo(update, context):
    """Echo the user message."""
    update.message.reply_text(update.message.text)

def error(update, context):
    """Log Errors caused by Updates."""
    logger.warning('Update "%s" caused error "%s"', update, context.error)

def main():
    """Start the bot."""
    # Create the Updater and pass it your bot's token.
    # Make sure to set use_context=True to use the new context based callbacks
    # Post version 12 this will no longer be necessary
    updater = Updater(TOKEN, use_context=True)

    # Get the dispatcher to register handlers
    dp = updater.dispatcher

    # on different commands - answer in Telegram
    dp.add_handler(CommandHandler("start", start))
    dp.add_handler(CommandHandler("help", help))

    # on noncommand i.e message - echo the message on Telegram
    dp.add_handler(MessageHandler(Filters.text, echo))

    # log all errors
    dp.add_error_handler(error)

    # Start the Bot
    updater.start_webhook(listen="0.0.0.0",
                          port=int(PORT),
                          url_path=TOKEN)
    updater.bot.setWebhook('https://vast-refuge-19289.herokuapp.com/' + TOKEN)

    # Run the bot until you press Ctrl-C or the process receives SIGINT,
    # SIGTERM or SIGABRT. This should be used most of the time, since
    # start_polling() is non-blocking and will stop the bot gracefully.
    updater.idle()

if __name__ == '__main__':
    main()

